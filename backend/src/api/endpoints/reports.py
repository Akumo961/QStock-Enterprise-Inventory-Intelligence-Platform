"""
Reports API endpoints.

Generates downloadable PDF, Excel, and CSV reports for QStock administrators.

Supported report types:
- user_activity       : per-user borrow/return activity in a date range
- inventory_status    : current snapshot of every item
- transaction_history : transaction history in a date range
- overdue_items       : currently overdue borrows
- usage_analytics     : usage summary and category borrowing breakdown
- popular_items       : items ranked by borrow count
"""

import csv
import io
from collections import defaultdict
from collections.abc import Callable
from datetime import UTC, datetime, timedelta
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from src.core.database import get_db
from src.core.security import get_current_admin_user
from src.models.item import Item
from src.models.review import Review
from src.models.transaction import Transaction, TransactionStatus
from src.models.user import User

router = APIRouter(tags=["Reports"])


VALID_REPORT_TYPES = {
    "inventory_status",
    "overdue_items",
    "popular_items",
    "transaction_history",
    "usage_analytics",
    "user_activity",
}

VALID_FORMATS = {
    "csv",
    "excel",
    "pdf",
}

CONTENT_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "excel": (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    ),
    "pdf": "application/pdf",
}

EXTENSIONS = {
    "csv": "csv",
    "excel": "xlsx",
    "pdf": "pdf",
}


# ============================================================================
# DATE RANGE
# ============================================================================


def _parse_date_range(
    start_date: str | None,
    end_date: str | None,
) -> tuple[datetime, datetime]:
    """
    Parse an inclusive calendar date range into UTC datetimes.

    The returned end datetime is exclusive.

    Example:
        start_date=2026-08-01
        end_date=2026-08-10

    becomes:

        start = 2026-08-01 00:00:00 UTC
        end   = 2026-08-11 00:00:00 UTC
    """
    now = datetime.now(UTC)

    try:
        if start_date:
            start = datetime.strptime(
                start_date,
                "%Y-%m-%d",
            ).replace(tzinfo=UTC)
        else:
            start = (now - timedelta(days=30)).replace(
                hour=0,
                minute=0,
                second=0,
                microsecond=0,
            )

        if end_date:
            end = (
                datetime.strptime(
                    end_date,
                    "%Y-%m-%d",
                ).replace(tzinfo=UTC)
                + timedelta(days=1)
            )
        else:
            end = now

    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Dates must be in YYYY-MM-DD format.",
        ) from exc

    if start >= end:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="start_date must be before end_date.",
        )

    return start, end


# ============================================================================
# FORMAT HELPERS
# ============================================================================


def _enum_value(value: Any) -> str:
    """Return an enum's underlying value or a safe string representation."""
    return str(getattr(value, "value", value))


def _format_datetime(value: datetime | None) -> str:
    """Format a datetime for human-readable reports."""
    if value is None:
        return "-"

    return value.strftime("%Y-%m-%d %H:%M")


def _format_date(value: datetime | None) -> str:
    """Format a datetime as a calendar date."""
    if value is None:
        return "-"

    return value.strftime("%Y-%m-%d")


def _date_range_label(
    start: datetime,
    end: datetime,
) -> str:
    """Create a human-readable inclusive date-range label."""
    effective_end = end - timedelta(microseconds=1)

    return (
        f"{start.strftime('%Y-%m-%d')} to "
        f"{effective_end.strftime('%Y-%m-%d')}"
    )


# ============================================================================
# DATA BUILDERS
# ============================================================================


def _build_user_activity(
    db: Session,
    start: datetime,
    end: datetime,
) -> tuple[list[str], list[list[str]]]:
    """
    Build per-user borrowing and return activity.

    Borrow counts are based on ``borrowed_at``.
    Return counts are based on ``returned_at``.

    This is important because a transaction may be borrowed before the
    selected reporting period and returned during it.
    """
    headers = [
        "User",
        "Email",
        "Department",
        "Borrows",
        "Returns",
        "Currently Borrowed",
        "Overdue",
    ]

    users = (
        db.query(User)
        .order_by(User.full_name.asc())
        .all()
    )

    borrow_transactions = (
        db.query(Transaction)
        .filter(
            Transaction.borrowed_at >= start,
            Transaction.borrowed_at < end,
        )
        .all()
    )

    return_transactions = (
        db.query(Transaction)
        .filter(
            Transaction.returned_at >= start,
            Transaction.returned_at < end,
        )
        .all()
    )

    borrows_by_user: dict[Any, list[Transaction]] = defaultdict(list)
    returns_by_user: dict[Any, list[Transaction]] = defaultdict(list)

    for transaction in borrow_transactions:
        borrows_by_user[transaction.user_id].append(transaction)

    for transaction in return_transactions:
        returns_by_user[transaction.user_id].append(transaction)

    now = datetime.now(UTC)

    rows: list[list[str]] = []

    for user in users:
        user_borrows = borrows_by_user.get(user.id, [])
        user_returns = returns_by_user.get(user.id, [])

        if not user_borrows and not user_returns:
            continue

        active_borrows = sum(
            transaction.status == TransactionStatus.BORROWED
            for transaction in user_borrows
        )

        overdue_items = sum(
            transaction.status == TransactionStatus.BORROWED
            and transaction.due_date is not None
            and transaction.due_date < now
            for transaction in user_borrows
        )

        rows.append(
            [
                user.full_name,
                user.email,
                user.department or "-",
                str(len(user_borrows)),
                str(len(user_returns)),
                str(active_borrows),
                str(overdue_items),
            ]
        )

    return headers, rows


def _build_inventory_status(
    db: Session,
    start: datetime,
    end: datetime,
) -> tuple[list[str], list[list[str]]]:
    """Build a current inventory snapshot."""
    del start, end

    headers = [
        "Item Code",
        "Name",
        "Category",
        "Status",
        "Quantity",
        "Available",
        "Location",
    ]

    items = (
        db.query(Item)
        .order_by(Item.name.asc())
        .all()
    )

    rows = [
        [
            item.item_code,
            item.name,
            _enum_value(item.category),
            _enum_value(item.status),
            str(item.quantity),
            str(item.available_quantity),
            item.location or "-",
        ]
        for item in items
    ]

    return headers, rows


def _build_transaction_history(
    db: Session,
    start: datetime,
    end: datetime,
) -> tuple[list[str], list[list[str]]]:
    """Build transaction history for the selected borrowing period."""
    headers = [
        "Item",
        "User",
        "Status",
        "Quantity",
        "Borrowed At",
        "Due Date",
        "Returned At",
    ]

    transactions = (
        db.query(Transaction, Item, User)
        .join(
            Item,
            Item.id == Transaction.item_id,
        )
        .join(
            User,
            User.id == Transaction.user_id,
        )
        .filter(
            Transaction.borrowed_at >= start,
            Transaction.borrowed_at < end,
        )
        .order_by(
            Transaction.borrowed_at.desc(),
            Transaction.id.desc(),
        )
        .all()
    )

    rows = [
        [
            item.name,
            user.full_name,
            _enum_value(transaction.status),
            str(transaction.quantity),
            _format_datetime(transaction.borrowed_at),
            _format_date(transaction.due_date),
            _format_datetime(transaction.returned_at),
        ]
        for transaction, item, user in transactions
    ]

    return headers, rows


def _build_overdue_items(
    db: Session,
    start: datetime,
    end: datetime,
) -> tuple[list[str], list[list[str]]]:
    """
    Build the currently overdue transaction report.

    This report intentionally ignores the selected date range because it
    represents the current operational state of overdue inventory.
    """
    del start, end

    headers = [
        "Item",
        "User",
        "Borrowed At",
        "Due Date",
        "Days Overdue",
    ]

    now = datetime.now(UTC)

    transactions = (
        db.query(Transaction, Item, User)
        .join(
            Item,
            Item.id == Transaction.item_id,
        )
        .join(
            User,
            User.id == Transaction.user_id,
        )
        .filter(
            Transaction.status == TransactionStatus.BORROWED,
            Transaction.due_date.is_not(None),
            Transaction.due_date < now,
        )
        .order_by(
            Transaction.due_date.asc(),
            Transaction.id.asc(),
        )
        .all()
    )

    rows = [
        [
            item.name,
            user.full_name,
            _format_date(transaction.borrowed_at),
            _format_date(transaction.due_date),
            str(
                max(
                    (now - transaction.due_date).days,
                    0,
                )
            ),
        ]
        for transaction, item, user in transactions
    ]

    return headers, rows


def _build_usage_analytics(
    db: Session,
    start: datetime,
    end: datetime,
) -> tuple[list[str], list[list[str]]]:
    """Build usage summary and category borrowing statistics."""
    headers = [
        "Metric",
        "Value",
    ]

    total_borrows = (
        db.query(func.count(Transaction.id))
        .filter(
            Transaction.borrowed_at >= start,
            Transaction.borrowed_at < end,
        )
        .scalar()
        or 0
    )

    total_returns = (
        db.query(func.count(Transaction.id))
        .filter(
            Transaction.returned_at >= start,
            Transaction.returned_at < end,
        )
        .scalar()
        or 0
    )

    total_reviews = (
        db.query(func.count(Review.id))
        .filter(
            Review.created_at >= start,
            Review.created_at < end,
        )
        .scalar()
        or 0
    )

    issue_reviews = (
        db.query(func.count(Review.id))
        .filter(
            Review.created_at >= start,
            Review.created_at < end,
            Review.has_issue.is_(True),
        )
        .scalar()
        or 0
    )

    return_rate = (
        round(
            (total_returns / total_borrows) * 100,
            1,
        )
        if total_borrows
        else 0.0
    )

    issue_rate = (
        round(
            (issue_reviews / total_reviews) * 100,
            1,
        )
        if total_reviews
        else 0.0
    )

    rows = [
        ["Total Borrows", str(total_borrows)],
        ["Total Returns", str(total_returns)],
        ["Return Rate (%)", str(return_rate)],
        ["Total Reviews", str(total_reviews)],
        ["Issue Reports", str(issue_reviews)],
        ["Issue Rate (%)", str(issue_rate)],
    ]

    category_counts = (
        db.query(
            Item.category,
            func.count(Transaction.id).label("borrow_count"),
        )
        .join(
            Transaction,
            Transaction.item_id == Item.id,
        )
        .filter(
            Transaction.borrowed_at >= start,
            Transaction.borrowed_at < end,
        )
        .group_by(Item.category)
        .order_by(func.count(Transaction.id).desc())
        .all()
    )

    for category, borrow_count in category_counts:
        rows.append(
            [
                f"Borrows in '{_enum_value(category)}'",
                str(borrow_count),
            ]
        )

    return headers, rows


def _build_popular_items(
    db: Session,
    start: datetime,
    end: datetime,
) -> tuple[list[str], list[list[str]]]:
    """Build an item popularity ranking for the selected period."""
    headers = [
        "Item",
        "Category",
        "Borrow Count",
    ]

    results = (
        db.query(
            Item.name,
            Item.category,
            func.count(Transaction.id).label("borrow_count"),
        )
        .join(
            Transaction,
            Transaction.item_id == Item.id,
        )
        .filter(
            Transaction.borrowed_at >= start,
            Transaction.borrowed_at < end,
        )
        .group_by(
            Item.id,
            Item.name,
            Item.category,
        )
        .order_by(
            func.count(Transaction.id).desc(),
            Item.name.asc(),
        )
        .limit(50)
        .all()
    )

    rows = [
        [
            name,
            _enum_value(category),
            str(borrow_count),
        ]
        for name, category, borrow_count in results
    ]

    return headers, rows


ReportBuilder = Callable[
    [Session, datetime, datetime],
    tuple[list[str], list[list[str]]],
]


BUILDERS: dict[str, ReportBuilder] = {
    "user_activity": _build_user_activity,
    "inventory_status": _build_inventory_status,
    "transaction_history": _build_transaction_history,
    "overdue_items": _build_overdue_items,
    "usage_analytics": _build_usage_analytics,
    "popular_items": _build_popular_items,
}


# ============================================================================
# RENDERERS
# ============================================================================


def _render_csv(
    headers: list[str],
    rows: list[list[str]],
) -> bytes:
    """Render report data as UTF-8 CSV with Excel-compatible BOM."""
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)

    writer.writerow(headers)
    writer.writerows(rows)

    return buffer.getvalue().encode("utf-8-sig")


def _render_excel(
    headers: list[str],
    rows: list[list[str]],
    title: str,
) -> bytes:
    """Render report data as an XLSX workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    worksheet = workbook.active

    if worksheet is None:
        raise RuntimeError(
            "Unable to create the Excel worksheet."
        )

    worksheet.title = title[:31] or "Report"

    header_fill = PatternFill(
        start_color="1B4332",
        end_color="1B4332",
        fill_type="solid",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )
        cell.fill = header_fill
        cell.font = header_font

    for row_index, row in enumerate(
        rows,
        start=2,
    ):
        for column_index, value in enumerate(
            row,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        values = [
            len(str(row[column_index - 1]))
            for row in rows
            if column_index - 1 < len(row)
        ]

        max_length = max(
            [len(str(header)), *values],
            default=len(str(header)),
        )

        column_letter = worksheet.cell(
            row=1,
            column=column_index,
        ).column_letter

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 4,
            50,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    buffer = io.BytesIO()
    workbook.save(buffer)

    return buffer.getvalue()


def _render_pdf(
    headers: list[str],
    rows: list[list[str]],
    title: str,
    date_range_label: str,
) -> bytes:
    """Render report data as a landscape PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buffer = io.BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()

    elements = [
        Paragraph(title, styles["Title"]),
        Paragraph(date_range_label, styles["Normal"]),
        Spacer(1, 0.3 * inch),
    ]

    if not rows:
        elements.append(
            Paragraph(
                "No data found for the selected period.",
                styles["Normal"],
            )
        )
    else:
        table_data = [
            headers,
            *[
                [str(cell) for cell in row]
                for row in rows
            ],
        ]

        table = Table(
            table_data,
            repeatRows=1,
        )

        table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1B4332"),
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white,
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey,
                    ),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [
                            colors.white,
                            colors.HexColor("#F1F8F4"),
                        ],
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE",
                    ),
                ]
            )
        )

        elements.append(table)

    document.build(elements)

    return buffer.getvalue()


# ============================================================================
# REPORT GENERATION ENDPOINT
# ============================================================================


@router.get("/generate")
async def generate_report(
    report_type: str = Query(...),
    format: str = Query(...),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user),
) -> StreamingResponse:
    """Generate and download an administrator report."""

    del current_admin

    if report_type not in VALID_REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                "Invalid report_type. Must be one of: "
                f"{', '.join(sorted(VALID_REPORT_TYPES))}"
            ),
        )

    report_format = format.lower()

    if report_format not in VALID_FORMATS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                "Invalid format. Must be one of: "
                f"{', '.join(sorted(VALID_FORMATS))}"
            ),
        )

    start, end = _parse_date_range(
        start_date,
        end_date,
    )

    builder = BUILDERS[report_type]

    try:
        headers, rows = builder(
            db,
            start,
            end,
        )
    except SQLAlchemyError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve report data.",
        ) from exc

    title = report_type.replace(
        "_",
        " ",
    ).title()

    date_label = _date_range_label(
        start,
        end,
    )

    try:
        if report_format == "csv":
            content = _render_csv(
                headers,
                rows,
            )
        elif report_format == "excel":
            content = _render_excel(
                headers,
                rows,
                title,
            )
        else:
            content = _render_pdf(
                headers,
                rows,
                title,
                date_label,
            )
    except (
        OSError,
        RuntimeError,
        TypeError,
        ValueError,
    ) as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=(
                f"Failed to render {report_format} report."
            ),
        ) from exc

    filename = (
        f"{report_type}_"
        f"{datetime.now(UTC).strftime('%Y%m%d')}."
        f"{EXTENSIONS[report_format]}"
    )

    return StreamingResponse(
        io.BytesIO(content),
        media_type=CONTENT_TYPES[report_format],
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"'
            )
        },
    )
